import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict


class Loan:
    def __init__(
            self,
            principal: float,
            annual_rate: float,
            years: int,
            payments_per_year: int = 12,
            loan_type: str = "equal_payment",  # "equal_payment" 或 "equal_principal"
            prepayment: Dict[int, float] = None,  # {期数: 提前还款金额}
            full_prepayment_at: int = None  # 第 N 期一次性提前结清
    ):
        self.principal = principal
        self.annual_rate = annual_rate
        self.years = years
        self.payments_per_year = payments_per_year
        self.loan_type = loan_type
        self.prepayment = prepayment or {}
        self.full_prepayment_at = full_prepayment_at

        self.period_rate = self.annual_rate / self.payments_per_year
        self.total_periods = self.years * self.payments_per_year

        if self.loan_type == "equal_payment":
            self.payment = self._calculate_equal_payment()
        else:
            self.payment = None

    def _calculate_equal_payment(self) -> float:
        """
        计算等额本息每期还款
        """
        i = self.period_rate
        N = self.total_periods
        A = (self.principal * i * (1 + i) ** N) / ((1 + i) ** N - 1)
        return A

    def generate_schedule(self) -> List[Dict]:
        """
        生成还款计划，支持等额本息、等额本金、提前还款、一次性结清
        """
        balance = self.principal
        schedule = []

        N = self.total_periods
        i = self.period_rate
        equal_principal = balance / N if self.loan_type == "equal_principal" else None

        for period in range(1, int(N) + 1):
            if balance <= 0:
                break

            interest = balance * i

            if self.loan_type == "equal_payment":
                principal_payment = self.payment - interest
                total_payment = self.payment
            else:
                principal_payment = equal_principal
                total_payment = principal_payment + interest

            # 检查是否提前部分还款
            extra_prepayment = self.prepayment.get(period, 0)
            principal_payment += extra_prepayment
            total_payment += extra_prepayment

            balance -= principal_payment
            balance = max(balance, 0)

            # 检查一次性提前结清
            if self.full_prepayment_at and period == self.full_prepayment_at and balance > 0:
                total_payment += balance
                principal_payment += balance
                balance = 0

            schedule.append({
                "Period": period,
                "Principal Payment": round(principal_payment, 2),
                "Interest Payment": round(interest, 2),
                "Total Payment": round(total_payment, 2),
                "Remaining Balance": round(balance, 2)
            })

            if balance <= 0:
                break

        return schedule

    def summary(self):
        schedule = self.generate_schedule()
        total_payment = sum(item['Total Payment'] for item in schedule)
        total_interest = sum(item['Interest Payment'] for item in schedule)

        # 原始贷款利息（不含提前还款）
        origin_interest = self._calculate_original_interest()

        print("\n========== Loan Summary ==========")
        print(f"贷款本金：{self.principal} 元")
        print(f"年利率：{self.annual_rate * 100}%")
        print(f"贷款年限：{self.years} 年（共 {int(self.total_periods)} 期）")
        print(f"还款方式：{'等额本息' if self.loan_type == 'equal_payment' else '等额本金'}")
        if self.loan_type == "equal_payment":
            print(f"每期还款：{self.payment:.2f} 元")
        print(f"总还款：{total_payment:.2f} 元")
        print(f"总利息：{total_interest:.2f} 元")
        if total_interest < origin_interest:
            saved = origin_interest - total_interest
            print(f"✅ 提前还款节省利息：{saved:.2f} 元")
        print("===================================")

    def _calculate_original_interest(self):
        """
        计算没有提前还款时的总利息
        """
        balance = self.principal
        i = self.period_rate
        N = self.total_periods
        total_interest = 0

        if self.loan_type == "equal_payment":
            A = self._calculate_equal_payment()
            for _ in range(int(N)):
                interest = balance * i
                principal_payment = A - interest
                balance -= principal_payment
                total_interest += interest
        else:
            principal_payment = balance / N
            for _ in range(int(N)):
                interest = balance * i
                balance -= principal_payment
                total_interest += interest

        return round(total_interest, 2)

    def print_schedule(self):
        schedule = self.generate_schedule()

        print(f"\n{'期数':<5} {'本金':>12} {'利息':>12} {'总还款':>12} {'剩余贷款':>15}")
        for item in schedule:
            print(f"{item['Period']:<5} {item['Principal Payment']:>12.2f} {item['Interest Payment']:>12.2f} "
                  f"{item['Total Payment']:>12.2f} {item['Remaining Balance']:>15.2f}")

    def export_to_excel(self, filename: str = "loan_schedule.xlsx"):
        """
        导出美化过的 Excel 文件
        """
        schedule = self.generate_schedule()
        df = pd.DataFrame(schedule)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Loan Schedule", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Loan Schedule"]

            # 添加摘要信息
            summary_info = [
                ["贷款本金", self.principal],
                ["年利率", f"{self.annual_rate * 100}%"],
                ["贷款年限", f"{self.years} 年"],
                ["还款方式", '等额本息' if self.loan_type == 'equal_payment' else '等额本金'],
                ["每期还款", self.payment if self.payment else "不同"],
                ["总还款", df['Total Payment'].sum()],
                ["总利息", df['Interest Payment'].sum()],
            ]

            if self.prepayment:
                summary_info.append(["提前还款", str(self.prepayment)])
            if self.full_prepayment_at:
                summary_info.append(["一次性提前结清", f"第 {self.full_prepayment_at} 期"])

            origin_interest = self._calculate_original_interest()
            current_interest = df['Interest Payment'].sum()
            if current_interest < origin_interest:
                summary_info.append(["节省利息", origin_interest - current_interest])

            summary_start = len(df) + 4
            for idx, (label, value) in enumerate(summary_info):
                worksheet.cell(row=summary_start + idx, column=1, value=label)
                worksheet.cell(row=summary_start + idx, column=2, value=value)

            # 美化表头
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=1,
                    max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

            # 自动列宽
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

            # 冻结表头
            worksheet.freeze_panes = "A2"

            print(f"\n✅ 已导出美化 Excel：{filename}")

    def plot(self):
        schedule = self.generate_schedule()
        df = pd.DataFrame(schedule)

        plt.figure(figsize=(10, 6))

        plt.plot(df['Period'], df['Remaining Balance'], label='Remaining Balance', color='blue', linewidth=2)
        plt.plot(df['Period'], df['Principal Payment'], label='Principal Payment', color='green', linestyle='--')
        plt.plot(df['Period'], df['Interest Payment'], label='Interest Payment', color='red', linestyle=':')

        plt.title("Loan Payment Breakdown Over Time")
        plt.xlabel("Period")
        plt.ylabel("Amount (元)")
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend()
        plt.tight_layout()
        plt.show()


if __name__ == "__main__":
    # 示例：等额本息 + 提前部分还款 + 一次性提前结清
    loan = Loan(
        principal=1000000,
        annual_rate=0.048,
        years=30,
        loan_type="equal_payment",
        prepayment={12: 50000, 24: 30000},  # 第12期提前5万，第24期提前3万
        full_prepayment_at=60  # 第60期一次性还清
    )
    loan.summary()
    loan.print_schedule()
    loan.export_to_excel("loan_schedule.xlsx")
    loan.plot()

